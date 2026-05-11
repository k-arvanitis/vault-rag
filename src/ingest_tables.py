"""Excel/CSV loading utilities — parse .xlsx / .csv into raw row lists.

Used by ingest_table_rows.py (the table → DuckDB + Qdrant ingestion path).
Merged cells are forward-filled; blank rows and obvious index/cover sheets are skipped.
"""

from __future__ import annotations

import csv
from pathlib import Path
from typing import Any

import openpyxl

SKIP_SHEET_KEYWORDS = ("index", "contents", "cover", "notes")


def _sanitize_cell(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, str):
        stripped = value.strip()
        return stripped if stripped else None
    return value


def _forward_fill(row: list[Any]) -> list[Any]:
    last = None
    out: list[Any] = []
    for value in row:
        v = _sanitize_cell(value)
        if v is not None:
            last = v
        out.append(last)
    return out


def load_sheets(file_path: str) -> dict[str, list[list[Any]]]:
    path = Path(file_path)
    suffix = path.suffix.lower()
    if suffix in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
        return _load_xlsx(path)
    if suffix == ".csv":
        return _load_csv(path)
    raise ValueError(f"Unsupported file type: {path.suffix}. Expected .xlsx or .csv")


def _load_xlsx(path: Path) -> dict[str, list[list[Any]]]:
    wb = openpyxl.load_workbook(path, data_only=True)
    sheets: dict[str, list[list[Any]]] = {}
    for name in wb.sheetnames:
        if any(keyword in name.lower() for keyword in SKIP_SHEET_KEYWORDS):
            print(f"  Skipping '{name}'")
            continue
        ws = wb[name]
        rows: list[list[Any]] = []
        for row in ws.iter_rows():
            values = [cell.value for cell in row]
            if not any(v is not None and str(v).strip() != "" for v in values):
                continue
            rows.append(_forward_fill(values))
        sheets[name] = rows
    return sheets


def _load_csv(path: Path) -> dict[str, list[list[Any]]]:
    rows: list[list[Any]] = []
    with path.open("r", encoding="latin-1", newline="") as handle:
        reader = csv.reader(handle)
        for row in reader:
            if not any(str(v).strip() != "" for v in row):
                continue
            normalized = [v if str(v).strip() != "" else None for v in row]
            rows.append(_forward_fill(normalized))
    return {path.stem: rows}
