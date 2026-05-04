"""
Smart Excel Preprocessing Pipeline using LLM for structure detection.
"""

from dataclasses import dataclass
from typing import List, Optional, Callable, Dict, Any
import json
import re

import openpyxl
import pandas as pd


@dataclass
class SheetMetadata:
    sheet_name: str
    source_file: str
    notes: List[str]
    table_description: str
    original_header_rows: List[int]
    units_row: Optional[int]
    row_count: int
    column_count: int


@dataclass
class SheetResult:
    df: pd.DataFrame
    metadata: SheetMetadata


def serialize_rows(rows: List[List[Any]], limit: int = 20) -> str:
    """Serialize the first `limit` rows for the LLM prompt."""
    lines = []
    for i, row in enumerate(rows[:limit]):
        row_subset = row[:20]  # Limit to first 20 columns to avoid token explosion
        str_row = []
        for cell in row_subset:
            if cell is None:
                str_row.append("")
            else:
                text = str(cell).strip()
                if len(text) > 50:
                    text = text[:47] + "..."
                str_row.append(text)
        lines.append(f"Row {i}: {str_row}")
        
    result = "\n".join(lines)
    # Hard cap the string length to guarantee we never blow out token limits
    if len(result) > 4000:
        result = result[:3997] + "..."
    return result


def _looks_numeric(val: str) -> bool:
    s = val.strip().replace(",", "")
    try:
        float(s)
        return True
    except ValueError:
        return False


def heuristic_fallback(rows: List[List[Any]]) -> Dict[str, Any]:
    """Fallback if LLM returns invalid JSON."""
    header_row = 0
    for i, row in enumerate(rows[:20]):
        non_empty = [c for c in row if c is not None and str(c).strip()]
        if not non_empty:
            continue
        
        string_cells = [c for c in non_empty if isinstance(c, str) and not _looks_numeric(str(c))]
        if len(string_cells) / len(non_empty) > 0.5 and len(string_cells) > 1:
            header_row = i
            break
            
    return {
        "metadata_rows": list(range(header_row)),
        "header_rows": [header_row],
        "units_row": None,
        "data_start_row": header_row + 1,
        "notes": [],
        "table_description": "Extracted via heuristics."
    }


def _parse_llm_json(response: str, rows: List[List[Any]]) -> Dict[str, Any]:
    """Parse JSON from LLM output, extracting from markdown blocks if needed."""
    raw = response.strip()
    match = re.search(r"```(?:json)?\s*(\{.*?\})\s*```", raw, flags=re.DOTALL | re.IGNORECASE)
    if match:
        raw = match.group(1)
    else:
        # Try to find { ... }
        start = raw.find("{")
        end = raw.rfind("}")
        if start != -1 and end != -1 and start < end:
            raw = raw[start:end+1]
            
    try:
        data = json.loads(raw)
        # Validate structure roughly
        required = ["metadata_rows", "header_rows", "data_start_row", "notes", "table_description"]
        if not all(k in data for k in required):
            raise ValueError("Missing required keys")
        return data
    except Exception as e:
        print(f"Failed to parse LLM JSON: {e}, falling back to heuristics")
        return heuristic_fallback(rows)


def collapse_headers(rows: List[List[Any]], header_row_indices: List[int]) -> List[str]:
    """Collapse multi-row headers into single column names."""
    if not header_row_indices:
        # fallback if empty
        return [f"Col_{i}" for i in range(len(rows[0] if rows else []))]
        
    num_cols = len(rows[header_row_indices[0]])
    headers = [[] for _ in range(num_cols)]
    
    for r_idx in header_row_indices:
        if r_idx >= len(rows):
            continue
        row = rows[r_idx]
        for c_idx in range(num_cols):
            val = str(row[c_idx]).strip() if c_idx < len(row) and row[c_idx] is not None else ""
            if val:
                headers[c_idx].append(val)
                
    final_headers = []
    seen = {}
    for c_idx, h_parts in enumerate(headers):
        if not h_parts:
            base_name = f"Unnamed_{c_idx}"
        else:
            base_name = " > ".join(h_parts)
            
        # Ensure uniqueness
        if base_name not in seen:
            seen[base_name] = 0
            final_headers.append(base_name)
        else:
            seen[base_name] += 1
            final_headers.append(f"{base_name}_{seen[base_name]}")
            
    return final_headers


def find_data_end_row(rows: List[List[Any]], data_start_row: int) -> int:
    """Scan from the bottom up to strip trailing empty rows and footer notes."""
    end_idx = len(rows)
    for i in range(len(rows) - 1, data_start_row - 1, -1):
        row = rows[i]
        non_empty = [c for c in row if c is not None and str(c).strip() != ""]
        if not non_empty:
            end_idx = i
            continue

        empty_count = len(row) - len(non_empty)
        sparse_ratio = empty_count / len(row) if row else 1.0
        is_single_text_note = (
            len(non_empty) == 1
            and isinstance(non_empty[0], str)
            and not _looks_numeric(non_empty[0])
        )
        # 0.75 catches wide-sheet notes without breaking dense data rows; the
        # single-text rule handles compact two-column footers such as "Note".
        is_sparse = sparse_ratio > 0.75 or is_single_text_note
        
        if is_sparse:
            end_idx = i
        else:
            break
            
    return end_idx


def process_sheet(sheet_name: str, source_file: str, ws: Any, llm_fn: Callable[[str], str]) -> Optional[SheetResult]:
    # Read all rows into a list of lists.
    # openpyxl handles merged cells by putting the value in the top-left cell, and None in the rest.
    all_rows = []
    for row in ws.iter_rows(values_only=True):
        all_rows.append(list(row))
        
    if len(all_rows) < 3:
        print(f"Skipping sheet '{sheet_name}' (too few rows: {len(all_rows)})")
        return None
        
    serialized = serialize_rows(all_rows, limit=20)
    
    prompt = f"""You are analyzing a raw spreadsheet table from file '{source_file}', sheet '{sheet_name}'.
    
Here are the first 20 rows:
{serialized}

Return ONLY a JSON object with this exact structure:
{{
  "metadata_rows": [list of 0-indexed integer row indices that are titles/subtitles/notes],
  "header_rows": [list of 0-indexed integer row indices that form the column headers],
  "units_row": integer or null (optional: row index for units/scale info),
  "data_start_row": integer (first actual data row index),
  "notes": [list of strings extracted from the metadata rows],
  "table_description": "one sentence describing what this table contains"
}}
"""

    response = llm_fn(prompt)
    struct = _parse_llm_json(response, all_rows)
    
    header_rows = struct.get("header_rows", [])
    data_start = struct.get("data_start_row")
    if data_start is None:
        data_start = max(header_rows) + 1 if header_rows else 0
        
    units_row = struct.get("units_row", None)
    
    final_headers = collapse_headers(all_rows, header_rows)
    
    end_row = find_data_end_row(all_rows, data_start)
    
    data_rows = all_rows[data_start:end_row]
    footer_rows = all_rows[end_row:]
    
    if not data_rows:
        print(f"Warning: Sheet '{sheet_name}' yielded no data rows. Skipping.")
        return None
        
    df = pd.DataFrame(data_rows, columns=final_headers)
    
    # Extract footnotes
    notes = struct.get("notes", [])
    for row in footer_rows:
        for cell in row:
            if isinstance(cell, str) and cell.strip():
                if not _looks_numeric(cell):
                    notes.append(cell.strip())
    
    meta = SheetMetadata(
        sheet_name=sheet_name,
        source_file=source_file,
        notes=notes,
        table_description=struct.get("table_description", ""),
        original_header_rows=header_rows,
        units_row=units_row,
        row_count=len(data_rows),
        column_count=len(final_headers)
    )
    
    return SheetResult(df=df, metadata=meta)


def process_csv_file(path: str) -> Dict[str, SheetResult]:
    """Load a CSV file directly into a SheetResult — no LLM processing needed."""
    import os
    source_file = os.path.basename(path)
    sheet_name = os.path.splitext(source_file)[0]
    for enc in ("utf-8", "latin-1", "cp1252"):
        try:
            df = pd.read_csv(path, encoding=enc, low_memory=False)
            break
        except UnicodeDecodeError:
            continue
    else:
        raise ValueError(f"Could not decode {path} with any supported encoding")
    df = df.dropna(how="all").reset_index(drop=True)
    meta = SheetMetadata(
        sheet_name=sheet_name,
        source_file=source_file,
        notes=[],
        table_description=f"CSV table with {len(df)} rows and columns: {', '.join(str(c) for c in df.columns)}",
        original_header_rows=[0],
        units_row=None,
        row_count=len(df),
        column_count=len(df.columns),
    )
    return {sheet_name: SheetResult(df=df, metadata=meta)}


def process_file(path: str, llm_fn: Callable[[str], str]) -> Dict[str, SheetResult]:
    """Process an Excel or CSV file into SheetResult objects."""
    import os
    if path.lower().endswith(".csv"):
        return process_csv_file(path)
    wb = openpyxl.load_workbook(path, data_only=True)
    results = {}
    source_file = os.path.basename(path)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        res = process_sheet(sheet_name, source_file, ws, llm_fn)
        if res:
            results[sheet_name] = res

    return results
